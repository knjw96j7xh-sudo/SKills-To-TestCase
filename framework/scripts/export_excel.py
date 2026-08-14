#!/usr/bin/env python3
"""
export_excel.py — 将测试用例 JSON 导出为带样式的 Excel (.xlsx)

用法：
    python3 export_excel.py <input_json> <output_xlsx>
    python3 export_excel.py <input_json> <output_xlsx> --author "张三"

JSON 格式（由 md_to_json / 导出链路生成）：
{
  "meta": {
    "project": "项目名称",
    "module":  "模块名称",
    "generated_at": "2026-06-01",
    "author": "可选，默认编写人"
  },
  "testcases": [ ... ]
}
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from testcase_common import (
    TYPE_ORDER,
    TYPE_PRIORITY,
    case_sort_key,
    ensure_package,
    load_json_robust,
)

ensure_package("openpyxl")
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ─── 颜色 & 样式常量 ───────────────────────────────────────────────────────────

C = {
    "brand": "1E3A5F",
    "brand_mid": "2C5F8A",
    "brand_light": "E8F1F8",
    "white": "FFFFFF",
    "text_dark": "1A1A2E",
    "text_mid": "4A4A6A",
    "正向_bg": "EAF7EA",
    "异常_bg": "FDECEA",
    "边界_bg": "FFF8E1",
    "并发_bg": "E3F2FD",
    "正向_head": "2E7D32",
    "异常_head": "C62828",
    "边界_head": "E65100",
    "并发_head": "1565C0",
    "正向_badge": "43A047",
    "异常_badge": "E53935",
    "边界_badge": "FB8C00",
    "并发_badge": "1E88E5",
    "P0_bg": "FFCDD2",
    "P1_bg": "FFE0B2",
    "P2_bg": "FFF9C4",
    "P3_bg": "E8F5E9",
    "status_fail": "FFCDD2",
    "status_block": "FFE0B2",
    "status_pass": "C8E6C9",
    "status_skip": "E0E0E0",
    "border_light": "D0D8E4",
    "border_group": "B0BEC5",
}

STATUS_OPTS = ["未执行", "通过", "失败", "阻塞", "跳过"]
PRIORITY_ORDER = ["P0", "P1", "P2", "P3"]
DEFAULT_STATUS = "未执行"
MIN_ROW_HEIGHT = 22
MAX_ROW_HEIGHT = 120
LINE_HEIGHT = 14
ROW_PADDING = 6

# 列定义：(列头, 宽度, 列key或None)
COLUMNS = [
    ("用例ID", 10, "id"),
    ("所属模块", 14, "module"),
    ("测试点", 26, "test_point"),
    ("前置条件", 20, "precondition"),
    ("操作步骤", 48, "steps"),
    ("预期结果", 36, "expected"),
    ("关联检查点", 12, "checkpoint"),
    ("场景类型", 9, "type"),
    ("优先级", 8, "_priority"),
    ("执行状态", 10, "_status"),
    ("编写人", 10, "_author"),
    ("执行人", 8, None),
    ("备注", 18, "remark"),
]

N_COLS = len(COLUMNS)
COL_KEY_INDEX = {key: idx for idx, (_, _, key) in enumerate(COLUMNS, start=1)}
STATUS_COL = COL_KEY_INDEX["_status"]
AUTHOR_COL = COL_KEY_INDEX["_author"]
STEPS_COL = COL_KEY_INDEX["steps"]
EXPECTED_COL = COL_KEY_INDEX["expected"]
PRECONDITION_COL = COL_KEY_INDEX["precondition"]
TEST_POINT_COL = COL_KEY_INDEX["test_point"]

STEP_ITEM_RE = re.compile(
    r"^\s*(?:"
    r"(?:\d+)[.、．)]\s*"
    r"|(?:[一二三四五六七八九十]+)[、.．]\s*"
    r"|(?:[-*•])\s+"
    r")"
)
PLACEHOLDER_AUTHOR_RE = re.compile(r"^\[.*\]$|^填写|^待填|^TODO", re.I)


# ─── 样式工厂 ──────────────────────────────────────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def border(color: str = "D0D8E4", style: str = "thin") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def font(bold=False, color="1A1A2E", size=10, italic=False) -> Font:
    return Font(
        bold=bold,
        color=color,
        name="微软雅黑",
        size=size,
        italic=italic,
    )


def align(h="left", v="top", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ─── 文本规范化 ────────────────────────────────────────────────────────────────

def normalize_multiline(text) -> str:
    """统一步骤/预期的编号列表展示：去多余空行、规范编号前缀。"""
    if text is None:
        return ""
    raw = str(text).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not raw:
        return ""

    # 支持把「1. a 2. b」挤在同一行时按编号切开（仅当能识别到编号序列）
    if "\n" not in raw and re.search(r"(?:^|\s)\d+[.、．)]\s*\S", raw):
        parts = re.split(r"(?=(?:^|\s)\d+[.、．)]\s*)", raw)
        candidates = [p.strip() for p in parts if p and p.strip()]
        if len(candidates) >= 2 and sum(
            1 for p in candidates if re.match(r"^\d+[.、．)]\s*", p)
        ) >= 2:
            raw = "\n".join(candidates)

    lines = []
    for line in raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^(\d+)[.、．)]\s*(.*)$", line)
        if m:
            body = m.group(2).strip()
            lines.append(f"{int(m.group(1))}. {body}" if body else f"{int(m.group(1))}.")
            continue
        m = re.match(r"^([一二三四五六七八九十]+)[、.．]\s*(.*)$", line)
        if m:
            body = m.group(2).strip()
            lines.append(f"{m.group(1)}、{body}" if body else f"{m.group(1)}、")
            continue
        m = re.match(r"^[-*•]\s+(.*)$", line)
        if m:
            body = m.group(1).strip()
            lines.append(f"- {body}" if body else "-")
            continue
        lines.append(line)
    return "\n".join(lines)


def _display_width(text: str) -> float:
    """估算展示宽度：中文约 2，ASCII 约 1。"""
    width = 0.0
    for ch in text:
        if ord(ch) > 127:
            width += 2.0
        else:
            width += 1.0
    return width


def estimate_wrapped_lines(text: str, col_width: float) -> int:
    if not text:
        return 1
    # Excel 列宽单位近似字符宽；中文按 2 计，预留一点边距
    usable = max(4.0, col_width * 0.95)
    total = 0
    for line in str(text).split("\n"):
        w = _display_width(line)
        total += max(1, int((w + usable - 0.01) // usable))
    return max(1, total)


def estimate_row_height(
    steps: str,
    expected: str,
    precondition: str = "",
    test_point: str = "",
) -> float:
    col_widths = {idx: width for idx, (_, width, _) in enumerate(COLUMNS, start=1)}
    lines = max(
        estimate_wrapped_lines(steps, col_widths[STEPS_COL]),
        estimate_wrapped_lines(expected, col_widths[EXPECTED_COL]),
        estimate_wrapped_lines(precondition, col_widths[PRECONDITION_COL]),
        estimate_wrapped_lines(test_point, col_widths[TEST_POINT_COL]),
        1,
    )
    height = lines * LINE_HEIGHT + ROW_PADDING
    return max(MIN_ROW_HEIGHT, min(MAX_ROW_HEIGHT, height))


def load_default_author(input_json: str, explicit: str | None = None) -> str:
    if explicit and explicit.strip():
        return explicit.strip()

    input_path = Path(input_json).resolve()
    # 向上查找 .testcase-assets/project.config.md
    for parent in [input_path.parent, *input_path.parents]:
        config = parent / "project.config.md"
        if not config.is_file() and (parent / ".testcase-assets" / "project.config.md").is_file():
            config = parent / ".testcase-assets" / "project.config.md"
        if not config.is_file() and parent.name == ".testcase-assets":
            config = parent / "project.config.md"
        # 也兼容 history 子目录向上两级
        if not config.is_file():
            candidate = parent / ".testcase-assets" / "project.config.md"
            if candidate.is_file():
                config = candidate
        if config.is_file():
            author = _parse_author_from_config(config)
            if author:
                return author
    return ""


def _parse_author_from_config(path: Path) -> str:
    try:
        text = path.read_text(encoding="utf-8")
    except OSError:
        return ""
    for line in text.splitlines():
        if "测试负责人" not in line:
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) >= 2 and "测试负责人" in cells[0]:
            value = cells[1].strip()
            if value and not PLACEHOLDER_AUTHOR_RE.search(value):
                return value
    return ""


# ─── 摘要行（第 1 行）──────────────────────────────────────────────────────────

def write_summary_row(ws, project: str, module: str, total: int, generated_at: str):
    seg = N_COLS // 4
    spans = [
        (1, seg, f"项目：{project or '-'}"),
        (seg + 1, seg * 2, f"模块：{module or '-'}"),
        (seg * 2 + 1, seg * 3, f"共 {total} 条用例"),
        (seg * 3 + 1, N_COLS, f"日期：{generated_at}"),
    ]
    for c1, c2, text in spans:
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        cell = ws.cell(row=1, column=c1, value=text)
        cell.font = Font(bold=True, color=C["white"], name="微软雅黑", size=10)
        cell.fill = fill(C["brand_mid"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = Border(
            left=Side(style="medium", color=C["white"]),
            right=Side(style="medium", color=C["white"]),
            top=Side(style="thin", color=C["brand"]),
            bottom=Side(style="thin", color=C["brand"]),
        )
    ws.row_dimensions[1].height = 28


# ─── 列名行（第 2 行）──────────────────────────────────────────────────────────

def write_header_row(ws):
    for col_idx, (name, _, _key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=name)
        cell.font = font(bold=True, color=C["white"], size=10)
        cell.fill = fill(C["brand"])
        cell.alignment = align(h="center", v="center")
        cell.border = Border(
            left=Side(style="thin", color=C["brand_mid"]),
            right=Side(style="thin", color=C["brand_mid"]),
            top=Side(style="medium", color=C["white"]),
            bottom=Side(style="medium", color=C["brand_light"]),
        )
    ws.row_dimensions[2].height = 24


# ─── 分组视觉分隔 ──────────────────────────────────────────────────────────────

def apply_group_separator(ws, row_idx: int, tc_type: str):
    sep_color = C.get(f"{tc_type}_head", C["brand"])
    for col_idx in range(1, N_COLS + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        existing = cell.border
        cell.border = Border(
            left=existing.left,
            right=existing.right,
            top=Side(style="medium", color=sep_color),
            bottom=existing.bottom,
        )


# ─── 用例数据行 ────────────────────────────────────────────────────────────────

def write_testcase_row(ws, row_idx: int, tc: dict, row_in_group: int, default_author: str):
    tc_type = tc.get("type", "正向")
    if row_in_group % 2 == 1:
        bg = C.get(f"{tc_type}_bg", "F9F9F9")
    else:
        bg = "FFFFFF"

    priority = tc.get("priority") or TYPE_PRIORITY.get(tc_type, "P2")
    p_bg = C.get(f"{priority}_bg", "FFFFFF")

    steps_display = normalize_multiline(tc.get("steps", ""))
    expected_display = normalize_multiline(tc.get("expected", ""))
    precondition_display = normalize_multiline(tc.get("precondition", ""))
    test_point_display = str(tc.get("test_point", "") or "").strip()

    raw_cp = tc.get("checkpoint", "")
    if isinstance(raw_cp, list):
        checkpoint_str = ", ".join(str(x) for x in raw_cp)
    else:
        checkpoint_str = str(raw_cp or "")

    author = str(tc.get("author") or default_author or "").strip()
    status = str(tc.get("status") or DEFAULT_STATUS).strip() or DEFAULT_STATUS
    if status not in STATUS_OPTS:
        status = DEFAULT_STATUS

    values = {
        "id": tc.get("id", ""),
        "module": tc.get("module", ""),
        "test_point": test_point_display,
        "precondition": precondition_display,
        "steps": steps_display,
        "expected": expected_display,
        "checkpoint": checkpoint_str,
        "type": tc_type,
        "_priority": priority,
        "_status": status,
        "_author": author,
        "remark": tc.get("remark", ""),
        None: "",
    }

    for col_idx, (_, _, key) in enumerate(COLUMNS, start=1):
        val = values.get(key, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=val)

        if key == "_priority":
            cell.fill = fill(p_bg)
            cell.font = font(bold=True, size=9, color=C["text_dark"])
        elif key == "type":
            badge_bg = C.get(f"{tc_type}_badge", C["brand"])
            cell.fill = fill(badge_bg)
            cell.font = Font(bold=True, color=C["white"], name="微软雅黑", size=9)
        elif key == "_status":
            cell.fill = fill(bg)
            cell.font = font(size=9, color=C["text_dark"])
        else:
            cell.fill = fill(bg)
            cell.font = font(size=10, color=C["text_dark"])

        if key in ("id", "checkpoint", "type", "_priority", "_status", None):
            h_align = "center"
        else:
            h_align = "left"

        cell.alignment = Alignment(horizontal=h_align, vertical="top", wrap_text=True)
        cell.border = border(C["border_light"])

    ws.row_dimensions[row_idx].height = estimate_row_height(
        steps_display,
        expected_display,
        precondition_display,
        test_point_display,
    )


def apply_status_validation(ws, first_data_row: int, last_data_row: int):
    if last_data_row < first_data_row:
        return
    col_letter = get_column_letter(STATUS_COL)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_OPTS) + '"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="执行状态",
        error="请选择：未执行 / 通过 / 失败 / 阻塞 / 跳过",
    )
    dv.add(f"{col_letter}{first_data_row}:{col_letter}{last_data_row}")
    ws.add_data_validation(dv)

    # 条件格式：失败/阻塞/通过/跳过 浅色提示（不覆盖整行，只标状态列）
    status_range = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
    rules = [
        (f'${col_letter}{first_data_row}="失败"', C["status_fail"]),
        (f'${col_letter}{first_data_row}="阻塞"', C["status_block"]),
        (f'${col_letter}{first_data_row}="通过"', C["status_pass"]),
        (f'${col_letter}{first_data_row}="跳过"', C["status_skip"]),
    ]
    for formula, color in rules:
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(formula=[formula], fill=fill(color)),
        )


# ─── 列宽 ──────────────────────────────────────────────────────────────────────

def set_column_widths(ws):
    for col_idx, (_, width, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ─── 统计 Sheet ────────────────────────────────────────────────────────────────

def _write_stat_block(ws, start_row: int, title: str, headers: list[str], rows: list[list]):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, color=C["white"], name="微软雅黑", size=11)
    title_cell.fill = fill(C["brand"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 24

    header_row = start_row + 1
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = font(bold=True, color=C["white"])
        cell.fill = fill(C["brand_mid"])
        cell.alignment = align(h="center", v="center")
        cell.border = border()
    ws.row_dimensions[header_row].height = 20

    data_start = header_row + 1
    for offset, row_values in enumerate(rows):
        r = data_start + offset
        for c, val in enumerate(row_values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border()
            cell.alignment = align(h="center" if c > 1 else "left", v="center", wrap=False)
            if c == 1 and isinstance(val, str) and val in (
                "正向", "异常", "边界", "并发", *PRIORITY_ORDER
            ):
                badge = C.get(f"{val}_badge") or C.get(f"{val}_bg") or C["brand_light"]
                if val in ("正向", "异常", "边界", "并发"):
                    cell.fill = fill(C.get(f"{val}_badge", C["brand"]))
                    cell.font = Font(bold=True, color=C["white"], name="微软雅黑", size=10)
                else:
                    cell.fill = fill(C.get(f"{val}_bg", C["brand_light"]))
                    cell.font = font(bold=True, size=10)
            else:
                cell.fill = fill(C["brand_light"] if offset % 2 else "FFFFFF")
                cell.font = font(size=10)
        ws.row_dimensions[r].height = 20

    return data_start + len(rows)


def write_stat_sheet(wb, testcases: list):
    ws = wb.create_sheet("统计")

    total = len(testcases) or 1
    type_counts = defaultdict(int)
    priority_counts = defaultdict(int)
    module_counts = defaultdict(int)

    for tc in testcases:
        t = tc.get("type", "正向") or "正向"
        type_counts[t] += 1
        p = tc.get("priority") or TYPE_PRIORITY.get(t, "P2")
        priority_counts[p] += 1
        module = str(tc.get("module") or "").strip() or "(未填模块)"
        module_counts[module] += 1

    type_rows = []
    for t in TYPE_ORDER:
        cnt = type_counts.get(t, 0)
        type_rows.append([t, cnt, f"{cnt / total * 100:.1f}%"])
    type_rows.append(["合计", len(testcases), "100%"])

    next_row = _write_stat_block(
        ws, 1, "用例场景分布统计", ["场景类型", "用例数量", "占比"], type_rows
    )

    # 柱状图（场景）
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "用例场景分布"
    chart.y_axis.title = "用例数量"
    chart.x_axis.title = "场景类型"
    chart.style = 10
    chart.width = 14
    chart.height = 10
    data_ref = Reference(ws, min_col=2, max_col=2, min_row=2, max_row=2 + len(TYPE_ORDER))
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=2 + len(TYPE_ORDER))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "1E3A5F"
    chart.series[0].graphicalProperties.line.solidFill = "1E3A5F"
    ws.add_chart(chart, "E2")

    next_row += 2
    priority_rows = []
    for p in PRIORITY_ORDER:
        cnt = priority_counts.get(p, 0)
        priority_rows.append([p, cnt, f"{cnt / total * 100:.1f}%"])
    priority_rows.append(["合计", len(testcases), "100%"])
    next_row = _write_stat_block(
        ws, next_row, "优先级分布", ["优先级", "用例数量", "占比"], priority_rows
    )

    next_row += 2
    # 模块：Top 12 + 其他
    sorted_modules = sorted(module_counts.items(), key=lambda x: (-x[1], x[0]))
    top_n = 12
    module_rows = []
    shown = 0
    for name, cnt in sorted_modules[:top_n]:
        module_rows.append([name, cnt, f"{cnt / total * 100:.1f}%"])
        shown += cnt
    other = len(testcases) - shown
    if other > 0:
        module_rows.append(["其他", other, f"{other / total * 100:.1f}%"])
    module_rows.append(["合计", len(testcases), "100%"])
    _write_stat_block(
        ws, next_row, "所属模块分布", ["所属模块", "用例数量", "占比"], module_rows
    )

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10


# ─── 打印设置 ──────────────────────────────────────────────────────────────────

def apply_print_settings(ws, total_rows: int):
    from openpyxl.worksheet.page import PageMargins

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3
    )
    ws.oddHeader.center.text = "&B测试用例清单"
    ws.oddFooter.right.text = "第 &P 页 / 共 &N 页"


# ─── 主函数 ────────────────────────────────────────────────────────────────────

def export_excel(input_json: str, output_xlsx: str, author: str | None = None):
    import os

    if not os.path.exists(input_json):
        print(f"[FAIL] 输入文件不存在: {input_json}")
        sys.exit(1)

    output_dir = os.path.dirname(output_xlsx)
    if output_dir and not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
        except Exception as e:
            print(f"[FAIL] 无法创建输出目录: {e}")
            sys.exit(1)

    try:
        data = load_json_robust(input_json)
    except ValueError as e:
        print(f"[FAIL] {e}")
        sys.exit(1)
    except UnicodeDecodeError:
        print(f"[FAIL] 文件编码错误，请确保为 UTF-8: {input_json}")
        sys.exit(1)
    except Exception as e:
        print(f"[FAIL] 读取文件失败: {e}")
        sys.exit(1)

    try:
        with open(input_json, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except (OSError, IOError):
        pass

    meta = data.get("meta", {})
    project = meta.get("project", "测试用例")
    module = meta.get("module", "")
    generated_at = meta.get("generated_at", datetime.now().strftime("%Y-%m-%d"))
    testcases = data.get("testcases", [])
    default_author = (
        (author or "").strip()
        or str(meta.get("author") or "").strip()
        or load_default_author(input_json)
    )

    if not isinstance(testcases, list):
        print(
            f"[FAIL] JSON 中 testcases 字段必须是数组，当前类型: {type(testcases).__name__}"
        )
        sys.exit(1)

    if not testcases:
        print("[WARN] testcases 为空，将生成仅含标题的 Excel 文件")

    testcases.sort(key=lambda tc: case_sort_key(tc.get("id", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    write_summary_row(ws, project, module, len(testcases), generated_at)
    write_header_row(ws)

    current_row = 3
    prev_type = None
    for global_idx, tc in enumerate(testcases, start=1):
        tc_type = tc.get("type", "正向")
        write_testcase_row(ws, current_row, tc, global_idx, default_author)
        if tc_type != prev_type:
            apply_group_separator(ws, current_row, tc_type)
            prev_type = tc_type
        current_row += 1

    last_data_row = current_row - 1
    first_data_row = 3

    ws.freeze_panes = "A3"
    if last_data_row >= 2:
        ws.auto_filter.ref = f"A2:{get_column_letter(N_COLS)}{last_data_row}"
    if last_data_row >= first_data_row:
        apply_status_validation(ws, first_data_row, last_data_row)

    set_column_widths(ws)
    apply_print_settings(ws, current_row)
    write_stat_sheet(wb, testcases)

    try:
        wb.save(output_xlsx)
    except PermissionError:
        print(f"[FAIL] 无写入权限: {output_xlsx}")
        sys.exit(1)
    except Exception as e:
        print(f"[FAIL] 写入 Excel 失败: {e}")
        sys.exit(1)

    print(f"[OK] Excel 已生成：{output_xlsx}（共 {len(testcases)} 条用例）")
    print("     包含列：用例ID / 所属模块 / 测试点 / 前置条件 / 操作步骤 / 预期结果")
    print("           关联检查点 / 场景类型 / 优先级 / 执行状态 / 编写人 / 执行人 / 备注")
    print("     执行状态已加下拉；统计表含场景 / 优先级 / 模块分布")
    if default_author:
        print(f"     默认编写人：{default_author}")


def main():
    parser = argparse.ArgumentParser(description="导出测试用例 Excel")
    parser.add_argument("input_json", help="export_data.json 路径")
    parser.add_argument("output_xlsx", help="输出 .xlsx 路径")
    parser.add_argument(
        "--author",
        default="",
        help="默认编写人（优先于 project.config / meta.author）",
    )
    args = parser.parse_args()
    export_excel(args.input_json, args.output_xlsx, author=args.author or None)


if __name__ == "__main__":
    main()
