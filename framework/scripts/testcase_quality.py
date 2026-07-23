#!/usr/bin/env python3
"""检查测试用例内容质量并生成交付审计摘要。"""

import argparse
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path


FIELDS = (
    "id",
    "module",
    "test_point",
    "precondition",
    "steps",
    "expected",
    "checkpoint",
    "type",
    "priority",
    "remark",
)
HARD_REQUIRED = ("id", "module", "test_point", "steps", "expected", "type", "priority")
HEADER_TO_FIELD = {
    "用例ID": "id",
    "所属模块": "module",
    "测试点": "test_point",
    "前置条件": "precondition",
    "操作步骤": "steps",
    "预期结果": "expected",
    "关联检查点": "checkpoint",
    "场景类型": "type",
    "优先级": "priority",
    "备注": "remark",
}
VALID_TYPES = {"正向", "异常", "边界", "并发"}
VALID_PRIORITIES = {"P0", "P1", "P2", "P3"}
FUZZY_PATTERNS = (
    (re.compile(r"正常(?:显示|处理|运行|返回|完成|保存|提交|跳转)?"), "“正常”缺少可验证标准"),
    (re.compile(r"正确(?:显示|返回|处理|保存|计算)?"), "“正确”缺少明确结果"),
    (re.compile(r"符合预期"), "“符合预期”未说明具体预期"),
    (re.compile(r"无异常"), "“无异常”未说明可观察结果"),
    (re.compile(r"适当"), "“适当”缺少量化标准"),
    (re.compile(r"相关提示"), "“相关提示”未给出提示内容"),
)
STEP_NUMBER_RE = re.compile(r"(?:^|[\n\r]|\s)(\d+)[.、]\s*")
FORMULA_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#N/A", "#NULL!")


@dataclass(frozen=True)
class Issue:
    severity: str
    code: str
    case_id: str
    field: str
    message: str


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def load_json_cases(path: Path) -> tuple[dict, list[dict]]:
    with path.open("r", encoding="utf-8-sig") as handle:
        data = json.load(handle)
    if isinstance(data, list):
        data = {"testcases": data}
    if not isinstance(data, dict) or not isinstance(data.get("testcases", []), list):
        raise ValueError("JSON 顶层必须是对象，且 testcases 必须是数组")
    return data.get("meta", {}), data.get("testcases", [])


def _split_markdown_row(line: str) -> list[str]:
    return [cell.strip().replace("<br>", "\n") for cell in line.strip().strip("|").split("|")]


def load_markdown_cases(path: Path) -> tuple[dict, list[dict]]:
    lines = path.read_text(encoding="utf-8-sig").splitlines()
    column_map: dict[int, str] = {}
    cases: list[dict] = []
    title = ""

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("## ") and not title:
            title = stripped[3:].strip()
        if stripped.startswith("|") and "用例ID" in stripped:
            headers = _split_markdown_row(stripped)
            column_map = {
                index: HEADER_TO_FIELD[header]
                for index, header in enumerate(headers)
                if header in HEADER_TO_FIELD
            }
            continue
        if not column_map or not stripped.startswith("|"):
            continue
        cells = _split_markdown_row(stripped)
        case_id_index = next((index for index, field in column_map.items() if field == "id"), None)
        if case_id_index is None or case_id_index >= len(cells):
            continue
        if not re.fullmatch(r"[A-Za-z]+-\d+[A-Za-z]*", cells[case_id_index]):
            continue
        case = {field: cells[index] if index < len(cells) else "" for index, field in column_map.items()}
        cases.append(case)

    if not column_map:
        raise ValueError("Markdown 中未找到包含“用例ID”的表头")
    return {"project": title, "module": title}, cases


def load_cases(path: Path) -> tuple[dict, list[dict]]:
    if path.suffix.lower() == ".json":
        return load_json_cases(path)
    if path.suffix.lower() == ".md":
        return load_markdown_cases(path)
    raise ValueError("仅支持 .json 或 .md 输入")


def _numbering_issues(case_id: str, steps: str, expected: str) -> list[Issue]:
    issues: list[Issue] = []
    step_numbers = [int(value) for value in STEP_NUMBER_RE.findall(f"\n{steps}")]
    expected_numbers = [int(value) for value in STEP_NUMBER_RE.findall(f"\n{expected}")]

    if len(step_numbers) != len(set(step_numbers)):
        issues.append(Issue("ERROR", "STEP_DUPLICATE", case_id, "steps", "操作步骤存在重复编号"))
    if step_numbers and step_numbers != list(range(1, len(step_numbers) + 1)):
        issues.append(Issue("WARN", "STEP_SEQUENCE", case_id, "steps", "操作步骤编号应从 1 开始并连续递增"))
    if len(expected_numbers) != len(set(expected_numbers)):
        issues.append(Issue("ERROR", "EXPECTED_DUPLICATE", case_id, "expected", "预期结果存在重复编号"))
    if step_numbers and expected_numbers and step_numbers != expected_numbers:
        issues.append(Issue("WARN", "STEP_EXPECTED_MISMATCH", case_id, "expected", "步骤编号与预期结果编号无法一一对应"))
    return issues


def _wording_issues(case_id: str, field: str, value: str) -> list[Issue]:
    issues: list[Issue] = []
    if re.search(r"(?<![A-Za-z])button(?![A-Za-z])", value, re.IGNORECASE):
        issues.append(Issue("WARN", "TERM_BUTTON", case_id, field, "请统一使用中文术语“按钮”"))
    for pattern, message in FUZZY_PATTERNS:
        if pattern.search(value):
            issues.append(Issue("WARN", "FUZZY_WORDING", case_id, field, message))
            break
    return issues


def _quote_issues(case_id: str, fields: list[tuple[str, str]]) -> list[Issue]:
    styles = set()
    unbalanced = []
    for field, value in fields:
        if '"' in value:
            styles.add("英文双引号")
            if value.count('"') % 2:
                unbalanced.append(field)
        if "“" in value or "”" in value:
            styles.add("中文双引号")
            if value.count("“") != value.count("”"):
                unbalanced.append(field)
        if "「" in value or "」" in value:
            styles.add("直角引号")
            if value.count("「") != value.count("」"):
                unbalanced.append(field)

    issues = []
    if len(styles) > 1:
        issues.append(Issue("WARN", "QUOTE_MIXED", case_id, "*", f"混用了{'、'.join(sorted(styles))}"))
    if unbalanced:
        issues.append(Issue("WARN", "QUOTE_UNBALANCED", case_id, ",".join(sorted(set(unbalanced))), "存在未成对引号"))
    return issues


def inspect_cases(cases: list[dict]) -> list[Issue]:
    issues: list[Issue] = []
    ids = [_text(case.get("id")) for case in cases]
    duplicate_ids = {case_id for case_id, count in Counter(ids).items() if case_id and count > 1}

    for index, case in enumerate(cases, start=1):
        case_id = _text(case.get("id")) or f"第{index}行"
        if case_id in duplicate_ids:
            issues.append(Issue("ERROR", "DUPLICATE_ID", case_id, "id", "用例 ID 重复"))
        for field in HARD_REQUIRED:
            if not _text(case.get(field)):
                issues.append(Issue("ERROR", "REQUIRED_EMPTY", case_id, field, "核心必填字段为空"))
        if not _text(case.get("checkpoint")):
            issues.append(Issue("WARN", "CHECKPOINT_EMPTY", case_id, "checkpoint", "未关联检查点"))

        case_type = _text(case.get("type"))
        priority = _text(case.get("priority"))
        if case_type and case_type not in VALID_TYPES:
            issues.append(Issue("ERROR", "INVALID_TYPE", case_id, "type", f"非法场景类型：{case_type}"))
        if priority and priority not in VALID_PRIORITIES:
            issues.append(Issue("ERROR", "INVALID_PRIORITY", case_id, "priority", f"非法优先级：{priority}"))

        steps = _text(case.get("steps"))
        expected = _text(case.get("expected"))
        issues.extend(_numbering_issues(case_id, steps, expected))
        text_fields = [(field, _text(case.get(field))) for field in ("test_point", "precondition", "steps", "expected")]
        for field, value in text_fields:
            issues.extend(_wording_issues(case_id, field, value))
        issues.extend(_quote_issues(case_id, text_fields))
    return issues


def inspect_workbook(path: Path) -> tuple[int, list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("检查 Excel 公式需要安装 requirements.lock 中的 openpyxl") from error

    formulas = 0
    errors: set[str] = set()
    workbook = load_workbook(path, data_only=False, read_only=True)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                    formulas += 1
                    if any(token in str(value).upper() for token in FORMULA_ERRORS):
                        errors.add(f"{sheet.title}!{cell.coordinate}: {value}")
                if cell.data_type == "e" or (isinstance(value, str) and value.upper() in FORMULA_ERRORS):
                    errors.add(f"{sheet.title}!{cell.coordinate}: {value}")
    workbook.close()
    return formulas, sorted(errors)


def _distribution(cases: list[dict], field: str, empty_label: str = "（空）") -> Counter:
    return Counter(_text(case.get(field)) or empty_label for case in cases)


def _table(counter: Counter, first_header: str) -> list[str]:
    lines = [f"| {first_header} | 数量 |", "|---|---:|"]
    lines.extend(f"| {name} | {count} |" for name, count in sorted(counter.items()))
    return lines


def render_audit(
    source: Path,
    cases: list[dict],
    issues: list[Issue],
    formula_count: int = 0,
    formula_errors: list[str] | None = None,
) -> str:
    formula_errors = formula_errors or []
    modules = _distribution(cases, "module")
    scenes = _distribution(cases, "type")
    blanks = Counter({field: sum(not _text(case.get(field)) for case in cases) for field in FIELDS})
    duplicates = sorted(case_id for case_id, count in _distribution(cases, "id").items() if case_id != "（空）" and count > 1)
    severity_counts = Counter(issue.severity for issue in issues)

    lines = [
        "# 测试用例审计摘要",
        "",
        f"> 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ",
        f"> 数据源：`{source}`",
        "",
        "## 总览",
        "",
        "| 指标 | 数量 |",
        "|---|---:|",
        f"| 用例 | {len(cases)} |",
        f"| 模块 | {len(modules)} |",
        f"| 错误 | {severity_counts.get('ERROR', 0)} |",
        f"| 警告 | {severity_counts.get('WARN', 0)} |",
        f"| 重复 ID | {len(duplicates)} |",
        f"| Excel 公式 | {formula_count} |",
        f"| 公式错误 | {len(formula_errors)} |",
        "",
        "## 模块分布",
        "",
        *_table(modules, "模块"),
        "",
        "## 场景分布",
        "",
        *_table(scenes, "场景类型"),
        "",
        "## 字段空值",
        "",
        *_table(blanks, "字段"),
        "",
        "## 重复 ID",
        "",
    ]
    lines.extend([f"- `{case_id}`" for case_id in duplicates] or ["- 无"])
    lines.extend(["", "## 内容质量与异常字段", ""])
    if issues:
        lines.extend(["| 级别 | 规则 | 用例 | 字段 | 说明 |", "|---|---|---|---|---|"])
        for issue in issues:
            message = issue.message.replace("|", "\\|")
            lines.append(f"| {issue.severity} | {issue.code} | {issue.case_id} | {issue.field} | {message} |")
    else:
        lines.append("- 未发现问题")
    lines.extend(["", "## Excel 公式检查", ""])
    if formula_errors:
        lines.extend(f"- `{item}`" for item in formula_errors)
    elif formula_count:
        lines.append(f"- 共检查 {formula_count} 个公式，未发现明确错误值或错误引用。")
    else:
        lines.append("- 未提供 Excel 文件或文件中没有公式。")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="export_data.json 或 2-用例定稿.md")
    parser.add_argument("--audit-output", type=Path, required=True, help="审计摘要输出路径")
    parser.add_argument("--xlsx", type=Path, help="可选的 Excel 文件，用于公式检查")
    parser.add_argument("--strict", action="store_true", help="发现 ERROR 级问题时返回退出码 1")
    args = parser.parse_args()

    try:
        _, cases = load_cases(args.input)
        issues = inspect_cases(cases)
        formula_count, formula_errors = (0, [])
        if args.xlsx:
            formula_count, formula_errors = inspect_workbook(args.xlsx)
    except (OSError, ValueError, json.JSONDecodeError, RuntimeError) as error:
        print(f"[FAIL] {error}")
        return 2

    args.audit_output.parent.mkdir(parents=True, exist_ok=True)
    args.audit_output.write_text(
        render_audit(args.input, cases, issues, formula_count, formula_errors),
        encoding="utf-8",
    )
    errors = sum(issue.severity == "ERROR" for issue in issues)
    warnings = sum(issue.severity == "WARN" for issue in issues)
    print(f"[AUDIT] {args.audit_output}（{len(cases)} 条，错误 {errors}，警告 {warnings}）")
    return 1 if args.strict and errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
